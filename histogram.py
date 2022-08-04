from openpyxl import load_workbook
import numpy as np
from matplotlib import pyplot
from matplotlib import colors

# Load workbook and sheet
wb = load_workbook(filename="Finger_Length_Data.xlsx", data_only=True)
ws = wb["Sheet1"]

# Grab 2-dimensional array
allCells = np.array([[cell.value for cell in row] for row in ws.iter_rows()])

# Separate into f(female) and m(male) 1-dimensional arrays
# Also convert from numpy to list and the elements from strings to floats.  UGH
f_header = allCells[0,0]
fdata = allCells[1:13,0]
fdata2 = fdata.tolist()
f_data = [float(element) for element in fdata2]

m_header = allCells[0,1]
mdata = allCells[1:13,1]
mdata2 = mdata.tolist()
m_data = [float(element) for element in mdata2]

# AND THEN!!!! WE HAVE TO RECOMBINE THEM!!!
all_header = [f_header, m_header]
all_data = [f_data, m_data]

# START THE PLOTTING
p = pyplot.figure()
p.set_figwidth(20)

colors=['pink', 'blue']
pyplot.hist(all_data, bins=np.arange(7.4, 9.0, .1), color=colors, label=all_header, rwidth=.97)
pyplot.legend(loc="upper right")
pyplot.xticks(np.arange(7.4, 9.0, .1))
pyplot.xlabel("Measurement /cm")
pyplot.yticks(range(0,7))
pyplot.ylabel("Frequency")
pyplot.title("Graph 1:  Left middle finger length measurements", fontweight="bold")
pyplot.show()